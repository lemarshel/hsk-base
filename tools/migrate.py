"""
One-time migration: index.html -> data/words.xlsx + data/groups.json + data/groups-data.js
Run from the project root:  python tools/migrate.py
Re-run only if the phonetic group structure itself changes.
"""
import json, re
from pathlib import Path
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT       = Path(__file__).parent.parent
HTML       = ROOT / "index.html"
XLSX       = ROOT / "data" / "words.xlsx"
GROUPS_J   = ROOT / "data" / "groups.json"
GROUPS_JS  = ROOT / "data" / "groups-data.js"

COLUMNS = [
    "id", "word", "pinyin", "en", "ru",
    "example_zh", "example_pinyin", "example_en", "example_ru",
    "pos", "phonetic_group", "component", "hsk",
]

def main():
    print(f"Reading {HTML} ...")
    soup = BeautifulSoup(HTML.read_text(encoding="utf-8"), "html.parser")

    groups          = []
    words           = []
    word_id         = 0
    seen_tbody      = set()
    pending_h2      = None
    pending_meta    = {"h3_inner": "", "h2": None}

    for tag in soup.body.descendants:
        if not hasattr(tag, "name"):
            continue

        # ── POS section heading (h2) ────────────────────────────────────────
        if tag.name == "h2" and "pos-group" in (tag.get("class") or []):
            zh_span  = tag.find("span", class_="pos-zh")
            zh_text  = zh_span.get_text(strip=True) if zh_span else ""
            ru_text  = tag.get_text(strip=True).replace(zh_text, "").strip()
            pending_h2 = {"id": tag.get("id", ""), "ru": ru_text, "zh": zh_text}

        # ── Phonetic group heading (h3) ─────────────────────────────────────
        elif tag.name == "h3" and "phonetic-group" in (tag.get("class") or []):
            pending_meta = {
                "h3_inner": tag.decode_contents().strip(),   # exact inner HTML
                "h2":       pending_h2,
            }
            pending_h2 = None

        # ── Word tbody ──────────────────────────────────────────────────────
        elif tag.name == "tbody" and tag.get("id") and tag["id"] not in seen_tbody:
            tbody_id = tag["id"]
            if tbody_id in ("learned-tbody", "fam-tbody"):
                continue
            seen_tbody.add(tbody_id)

            first_tr     = tag.find("tr", attrs={"data-key": True})
            component    = (first_tr or {}).get("data-component", "")
            component_py = (first_tr or {}).get("data-component-py", "")
            radical      = (first_tr or {}).get("data-radical", "")
            radical_py   = (first_tr or {}).get("data-radical-py", "")

            groups.append({
                "id":                  tbody_id,
                "h3_inner":            pending_meta["h3_inner"],
                "pos_heading_before":  pending_meta["h2"],
                "component":           component,
                "component_py":        component_py,
                "radical":             radical,
                "radical_py":          radical_py,
            })
            pending_meta = {"h3_inner": "", "h2": None}

            for tr in tag.find_all("tr", attrs={"data-key": True}):
                word_id += 1
                ex_zh = tr.find(class_="ex-zh")
                ex_py = tr.find(class_="ex-py")
                words.append({
                    "id":             word_id,
                    "word":           tr.get("data-key", ""),
                    "pinyin":         tr.get("data-py", ""),
                    "en":             tr.get("data-en", ""),
                    "ru":             tr.get("data-ru", ""),
                    "example_zh":     ex_zh.get_text(strip=True) if ex_zh else "",
                    "example_pinyin": ex_py.get_text(strip=True) if ex_py else "",
                    "example_en":     "",
                    "example_ru":     "",
                    "pos":            tr.get("data-section", ""),
                    "phonetic_group": tbody_id,
                    "component":      tr.get("data-component", ""),
                    "hsk":            int(tr.get("data-hsk") or 0),
                })

    print(f"  Groups extracted : {len(groups)}")
    print(f"  Words  extracted : {len(words)}")

    # ── Write groups.json ───────────────────────────────────────────────────
    GROUPS_J.write_text(json.dumps(groups, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {GROUPS_J}")

    # ── Write groups-data.js ────────────────────────────────────────────────
    GROUPS_JS.write_text(
        "window.HSK_GROUPS=" + json.dumps(groups, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {GROUPS_JS}")

    # ── Write words.xlsx ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Words"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2E5090")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_widths = {
        "id": 5, "word": 12, "pinyin": 16, "en": 28, "ru": 28,
        "example_zh": 40, "example_pinyin": 40, "example_en": 36, "example_ru": 36,
        "pos": 14, "phonetic_group": 14, "component": 12, "hsk": 5,
    }
    for i, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = col_widths.get(col, 14)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    da = Alignment(vertical="top", wrap_text=False)
    for w in words:
        ws.append([w[c] for c in COLUMNS])
        for cell in ws[ws.max_row]:
            cell.alignment = da

    wb.save(XLSX)
    print(f"Wrote {XLSX}  ({word_id} rows)")

if __name__ == "__main__":
    main()
